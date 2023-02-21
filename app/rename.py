import os


from tkinter import filedialog, Tk, ttk, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.cell import Cell
from app import dialog
from threading import Thread


class Rename:

    def __init__(self, app: Tk) -> None:
        self.app = app
        self.create_ui()

    def create_ui(self):
        open_button = ttk.Button(self.app, text='Rename',
                                 command=self.handle_open, style='BW.TLabel')
        open_button.place(x=150, y=630, width=200, height=50)

        self.row_input = ttk.Entry(self.app)
        self.row_input.place(x=400, y=630, width=200, height=30)
        self.row_input.insert(0, "Starting Row")

        self.col_input = ttk.Entry(self.app)
        self.col_input.place(x=500, y=630, width=100, height=30)
        self.col_input.insert(0, "Ending Column")

        self.end_input = ttk.Entry(self.app)
        self.end_input.place(x=600, y=630, width=100, height=30)
        self.end_input.insert(0, "Ending Row")

    def handle_open(self):
        thread = Thread(target=self.handle_rename)
        thread.start()

    def handle_rename(self):
        path = filedialog.askopenfilename()
        workbook = load_workbook(path)
        workspace = workbook.active

        renamed_workbook = Workbook()
        renamed_workspace = renamed_workbook.active

        count: int = 1
        for row in workspace.iter_rows():
            cell: Cell
            for cell in row:
                if cell.row < int(self.row_input.get()):
                    print("Breaking To Outer....")
                    break

                if cell.column_letter != self.col_input.get().upper and cell.row < int(self.end_input.get()):
                    renamed_workspace[f"{cell.column_letter}{cell.row}"]

        renamed_workspace.save("renamed.xlsx")
        print(f"Final row {count}")

        dialog.success()
