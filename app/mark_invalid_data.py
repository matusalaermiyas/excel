import os


from tkinter import filedialog, Tk, ttk, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.cell import Cell


class MakeRowRedWithInvalidData:
    """
        Make An Excel Row Containing Invalid Data Like "Empty, NG, Or I" Red
    """

    def __init__(self, app: Tk) -> None:
        self.app = app
        self.create_ui()

    def get_column_input(self) -> str:
        return self.column_input.get()

    def get_row_input(self) -> str:
        return self.row_input.get()

    def create_ui(self):

        label = ttk.Label(self.app, text="Max Column")
        label.place(x=40, y=40)

        self.column_input = ttk.Entry(self.app)
        self.column_input.place(x=150, y=40, width=300, height=30)
        self.column_input.insert(0, "s")

        row_label = ttk.Label(self.app, text="Starting Row")
        row_label.place(x=40, y=100)

        self.row_input = ttk.Entry(self.app)
        self.row_input.place(x=150, y=100, width=300, height=30)
        self.row_input.insert(0, "5")

        open_button = ttk.Button(self.app, text='Open Excel',
                                 command=self.handle_open, style='BW.TLabel')
        open_button.place(x=150, y=150, width=200, height=50)

    def handle_open(self):

        path = filedialog.askopenfilename()

        if path == "":
            messagebox.showerror(
                "Error", message="Make Sure To Select A File")
            return

        file_name = os.path.basename(path)

        pattern = PatternFill(start_color='FF0000',
                              end_color='FF0000', fill_type="solid")

        workbook: Workbook = load_workbook(path)
        workspace = workbook.active

        for row in workspace.iter_rows():
            make_it_red = False

            cell: Cell

            for cell in row:
                if cell.column_letter == self.get_column_input().upper() or cell.row < int(self.get_row_input()):
                    break

                if cell.value == "I" or cell.value == "NG" or cell.value == None:
                    make_it_red = True

            if make_it_red:
                for cell in row:
                    cell.fill = pattern

        workbook.save(filename=file_name)

        messagebox.showinfo("Success", "Task Finished Successfully")
