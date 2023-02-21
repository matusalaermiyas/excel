import os

from app import dialog, utils
from tkinter import filedialog, Tk, ttk
from openpyxl import load_workbook, Workbook


class CreateExcelWithInvalidIds:
    """
        Create Excel Doument Containing Invalid IDS
    """

    def __init__(self, app: Tk) -> None:
        self.app = app
        self.button_label = "Open Excel With Red Row And Get Ids"
        self.create_ui()

    def create_ui(self):
        get_invalid_id_button = ttk.Button(
            self.app, text=self.button_label, command=self.handle_get_invalid_id, style='BW.TLabel')
        get_invalid_id_button.place(x=150, y=250, height=50, width=400)

    def handle_get_invalid_id(self):
        path = filedialog.askopenfilename()

        if path == "":
            dialog.error()
            return

        workbook: Workbook = load_workbook(path)
        workspace = workbook.active

        invalid_workbook = Workbook()
        invalid_workspace = invalid_workbook.active
        file_name = os.path.basename(path)

        count: int = 1

        for row in workspace.iter_rows():

            for cell in row:
                if cell.fill.start_color.index != "00000000":
                    cell_name = f"A{count}"

                    invalid_workspace[cell_name] = cell.value
                    count = count + 1
                    break

        id_path = utils.join_from_cwd("invalid-ids")

        if not os.path.exists(id_path):
            os.makedirs(id_path)

        invalid_workbook.save(utils.join(id_path, file_name))

        dialog.success("Invalid IDS Created Successfully")
