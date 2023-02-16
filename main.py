from tkinter import filedialog, Tk, ttk, messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.cell import Cell

import os


try:
    app = Tk()

    x = ttk.Style()
    x.configure("BW.TLabel", foreground="white",
                background="darkgrey", font=('calibri', 15, 'bold'), padding=10)

    def handle_get_invalid_id():
        path = filedialog.askopenfilename()

        if path == "":
            messagebox.showerror(
                "Error", message="Make Sure To Select A File")
            return

        workbook: Workbook = load_workbook(path)
        workspace = workbook.active

        invalid_workbook = Workbook()
        invalid_workspace = invalid_workbook.active
        file_name = os.path.basename(path)

        count: int = 1

        for row in workspace.iter_rows():

            cell: Cell

            for cell in row:
                if cell.fill.start_color.index == "00FF0000":
                    cell_name = f"A{count}"
                    print(cell_name)
                    invalid_workspace[cell_name] = cell.value
                    count = count + 1
                    break

            invalid_workbook.save(f"IDS\Invalid IDS {file_name}")

        messagebox.showinfo("Success", "Invalid IDS created successfully")

    def get_column_input() -> str:
        return column_input.get()

    def get_row_input() -> str:
        return row_input.get()

    def handle_open():

        path = filedialog.askopenfilename()

        if path == "":
            messagebox.showerror(
                "Error", message="Make Sure To Select A File")
            return

        file_name = os.path.basename(path)

        pattern = PatternFill(start_color='FF0000',
                              end_color='FF0000', fill_type="solid")

        workbook: Workbook = load_workbook(path)
        workspace = workbook.active

        for row in workspace.iter_rows():
            make_it_red = False

            cell: Cell

            for cell in row:
                if cell.column_letter == get_column_input().upper() or cell.row < int(get_row_input()):
                    break

                if cell.value == "I" or cell.value == "NG" or cell.value == None:
                    make_it_red = True

            if make_it_red:
                for cell in row:
                    cell.fill = pattern

        workbook.save(filename=file_name)

        messagebox.showinfo("Success", "Task Finished Successfully")

    label = ttk.Label(app, text="Max Column")
    label.place(x=40, y=40)

    column_input = ttk.Entry(app,)
    column_input.place(x=150, y=40, width=200, height=30)
    column_input.insert(0, "Max Column")

    row_label = ttk.Label(app, text="Starting Row")
    row_label.place(x=40, y=100)

    row_input = ttk.Entry(app)
    row_input.place(x=150, y=100, width=200, height=30)
    row_input.insert(0, "Starting Row")

    open_button = ttk.Button(app, text='Open Excel',
                             command=handle_open, style='BW.TLabel')
    open_button.place(x=150, y=150, width=200, height=50)

    get_invalid_id_button = ttk.Button(
        app, text="Get invalid ", command=handle_get_invalid_id, style='BW.TLabel')
    get_invalid_id_button.place(x=150, y=220, height=50, width=200)

    app.geometry("500x500")
    app.title("Excel")
    app.mainloop()

except:
    messagebox.showerror("Error", "Error occured try again")
