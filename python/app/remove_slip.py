import os

from app import dialog, utils
from docx2pdf import convert
from openpyxl import load_workbook, Workbook
from openpyxl.cell import Cell
from pypdf import PdfReader, PdfWriter
from pypdf._page import PageObject
from tkinter import filedialog, Tk, ttk, messagebox


class RemoveARowWithInvalidId:
    """
        Remove A Page From Word Document Containing Invalid Slip
    """

    def __init__(self, app: Tk) -> None:
        self.app = app
        self.create_ui()

    def create_ui(self):
        row_label = ttk.Label(
            self.app, text="****************** Extract row with invalid ID *****************")
        row_label.place(x=150, y=350, height=50, width=400)
        open_excel_id_button = ttk.Button(
            self.app, text="Open Excel With Invalid Ids", command=self.handle_excel_open, style='BW.TLabel')
        open_excel_id_button.place(x=150, y=400, height=50, width=400)

        open_word_with_slip = ttk.Button(
            self.app, text="Open Word With Slip", command=self.handle_word_open, style='BW.TLabel')

        open_word_with_slip.place(x=150, y=470, height=50, width=400)

    def handle_excel_open(self):
        path = filedialog.askopenfilename()

        if path == "":
            dialog.error()
            return

        self.excel_path = path

    def handle_word_open(self):
        path = filedialog.askopenfilename()

        if path == "":

            dialog.error()

            return

        self.word_path = path

        self.start_remove()

    def start_remove(self):

        file_name = os.path.basename(self.excel_path)

        file_name_only, _ = os.path.splitext(file_name)

        temp_pdfs_path = utils.join_from_cwd("temp-pdfs")

        if not os.path.exists(temp_pdfs_path):
            os.makedirs(temp_pdfs_path)

        temp_pdf_path = utils.join(temp_pdfs_path, f"{file_name_only}.pdf")

        convert(self.word_path, temp_pdf_path)

        self.writer = PdfWriter()
        self.reader = PdfReader(temp_pdf_path)

        workbook: Workbook = load_workbook(self.excel_path)
        workspace = workbook.active

        first = True

        for row in workspace.iter_rows():

            cell: Cell

            for cell in row:

                if first:
                    self.filter(value=cell.value, first=True)
                    first = False

                else:
                    self.filter(value=cell.value)

                break

        invalid_slip_removed_path = os.path.join(
            os.getcwd(), "slip-removed-documents")
        if not os.path.exists(invalid_slip_removed_path):
            os.makedirs(invalid_slip_removed_path)

        self.writer.write(utils.join(
            invalid_slip_removed_path, f"{file_name_only}.pdf"))

        self.writer.close()

        os.unlink(temp_pdf_path)

        os.removedirs(temp_pdfs_path)

        dialog.success()

    def filter(self, value, first=False):
        if first:
            for page in self.reader.pages:

                text = str(page.extract_text())

                id_to_find = str(value)

                result = text.find(id_to_find)

                if result == -1:
                    self.writer.add_page(page)
        else:
            filtered: list(PageObject) = []

            for page in self.writer.pages:

                text = str(page.extract_text())

                id_to_find = str(value)

                result = text.find(id_to_find)

                if result == -1:
                    filtered.append(page)

            temp_writer: PdfWriter = PdfWriter()

            for f in filtered:
                temp_writer.add_page(f)

            self.writer = temp_writer
